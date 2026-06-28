"""Tool set for the CAMEL-style linear-pipeline MAS.

Self-contained (this MAS lives in its own folder, mirroring autogen_gc/). Three
plain callables + their OpenAI function schemas, a name->callable dispatch, and
per-benchmark TOOL_PROFILES. The agent inner loop (agent.py) passes the schemas
to the model as `tools=` and dispatches structured tool_calls back through here.

web_search/fetch_url reuse the Perplexity /search path; run_python executes in a
subprocess (same local-trust model as the autogen_gc harness). Tools are a
PER-BENCHMARK PROFILE, never gated on "closed vs open book": an empty profile is
just the degenerate case (pure reasoning), not a default.
"""
import json, os, subprocess, sys

import requests

PPLX_SEARCH_URL = os.environ.get("PPLX_SEARCH_URL", "https://api.perplexity.ai/search")
MAX_RESULTS = int(os.environ.get("WEBSURFER_MAX_RESULTS", "10"))
FETCH_MAX_CHARS = int(os.environ.get("FETCH_MAX_CHARS", "8000"))
READ_FILE_MAX_CHARS = int(os.environ.get("READ_FILE_MAX_CHARS", "40000"))
PYTHON_TIMEOUT = int(os.environ.get("RUN_PYTHON_TIMEOUT", "30"))

_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

_key = None


def _load_key():
    """PERPLEXITY_API_KEY from env, else from the nearest ancestor .env."""
    if os.environ.get("PERPLEXITY_API_KEY"):
        return os.environ["PERPLEXITY_API_KEY"]
    d = os.path.dirname(os.path.abspath(__file__))
    for _ in range(8):
        p = os.path.join(d, ".env")
        if os.path.exists(p):
            for line in open(p):
                if "=" in line and not line.lstrip().startswith("#"):
                    k, v = line.strip().split("=", 1)
                    os.environ.setdefault(k, v)
            break
        nd = os.path.dirname(d)
        if nd == d:
            break
        d = nd
    return os.environ.get("PERPLEXITY_API_KEY")


def web_search(query: str) -> str:
    """Search the web and return a numbered list of results (title, URL, snippet).

    Use this to find pages relevant to the task. The snippet is only a preview —
    to read a page's full content, pass its URL to fetch_url.
    """
    global _key
    if _key is None:
        _key = _load_key()
    if not _key:
        return "ERROR: PERPLEXITY_API_KEY not configured; web_search is unavailable."
    try:
        r = requests.post(
            PPLX_SEARCH_URL,
            headers={"Authorization": f"Bearer {_key}", "Content-Type": "application/json"},
            json={"query": query, "max_results": MAX_RESULTS}, timeout=60)
        r.raise_for_status()
        results = r.json().get("results", [])
    except Exception as e:  # surface to the model, don't crash the turn
        return f"ERROR: web_search failed for {query!r}: {e!r}"
    if not results:
        return f'No results for "{query}".'
    lines = [f'Search results for "{query}":', ""]
    for i, x in enumerate(results, 1):
        title = (x.get("title") or x.get("url") or "(untitled)").strip()
        url = (x.get("url") or "").strip()
        snippet = " ".join((x.get("snippet") or "").split())
        lines += [f"[{i}] {title}", f"    {url}"] + ([f"    {snippet}"] if snippet else []) + [""]
    return "\n".join(lines).rstrip()


# Structural boilerplate to drop before extracting text — nav/sidebar/footer/forms
# etc. are page chrome, not content. Removing them keeps the main text, so the
# head-truncation downstream cuts far less signal (it kept top-of-page chrome before).
_BOILERPLATE_TAGS = ["script", "style", "noscript", "svg", "nav", "header", "footer",
                     "aside", "form", "button", "figure", "iframe", "menu"]
_BOILERPLATE_ROLES = "[role=navigation],[role=banner],[role=contentinfo],[role=search]"


def fetch_url(url: str) -> str:
    """Fetch a web page and return its MAIN readable text (boilerplate removed).

    Extraction only (no chunking/summarizing): strip chrome tags, prefer the page's
    <main>/<article> content region if it marks one, then take the text. Use this
    after web_search to read the full text of a promising result.
    """
    try:
        from bs4 import BeautifulSoup
    except Exception as e:
        return f"ERROR: bs4 not available: {e!r}"
    try:
        r = requests.get(url, headers={"User-Agent": _UA}, timeout=45)
        r.raise_for_status()
    except Exception as e:
        return f"ERROR: fetch_url failed for {url!r}: {e!r}"
    ctype = r.headers.get("Content-Type", "").lower()
    is_pdf = "application/pdf" in ctype or url.lower().split("?")[0].endswith(".pdf")
    if is_pdf:                                       # PDFs on the web (GAIA "PDF access" tasks)
        text = _pdf_to_text(r.content)
    elif "html" not in ctype and "xml" not in ctype and ctype:
        text = r.text
    else:
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(_BOILERPLATE_TAGS):
            tag.decompose()
        try:
            for el in soup.select(_BOILERPLATE_ROLES):   # ARIA-marked chrome
                el.decompose()
        except Exception:
            pass
        # prefer the marked main-content region; fall back to body, then whole doc
        root = (soup.find("main") or soup.find("article")
                or soup.find(attrs={"role": "main"}) or soup.body or soup)
        text = "\n".join(l.strip() for l in root.get_text("\n").splitlines() if l.strip())
    if len(text) > FETCH_MAX_CHARS:
        text = text[:FETCH_MAX_CHARS] + f"\n\n[...truncated at {FETCH_MAX_CHARS} chars...]"
    return text or f"(no extractable text at {url})"


def run_python(code: str) -> str:
    """Execute a Python 3 snippet and return its stdout (and stderr on error).

    Use this for reliable computation — arithmetic, counting, parsing, dates, etc.
    Runs in a fresh subprocess; print() what you want to see. No state is kept
    between calls, so each call must be self-contained.
    """
    try:
        p = subprocess.run([sys.executable, "-c", code], capture_output=True,
                           text=True, timeout=PYTHON_TIMEOUT)
    except subprocess.TimeoutExpired:
        return f"ERROR: code timed out after {PYTHON_TIMEOUT}s."
    out = p.stdout or ""
    if p.returncode != 0:
        return (out + ("\n" if out else "") + "STDERR:\n" + (p.stderr or "")).strip() \
            or f"(exited {p.returncode} with no output)"
    return out.strip() or "(ran successfully, no stdout)"


def _pdf_to_text(data: bytes) -> str:
    """Extract text from PDF bytes (used by fetch_url for web PDFs and read_file)."""
    try:
        from pypdf import PdfReader
    except Exception as e:
        return f"ERROR: pypdf not available for PDF extraction: {e!r}"
    import io
    try:
        reader = PdfReader(io.BytesIO(data))
        parts = [p.extract_text() or "" for p in reader.pages]
        return "\n\n".join(t for t in parts if t.strip()).strip() or "(no extractable text in PDF)"
    except Exception as e:
        return f"ERROR: PDF parse failed: {e!r}"


def _xlsx_to_text(path: str) -> str:
    """Render an .xlsx workbook as tab-separated rows, one block per sheet."""
    try:
        from openpyxl import load_workbook
    except Exception as e:
        return f"ERROR: openpyxl not available for .xlsx: {e!r}"
    wb = load_workbook(path, read_only=True, data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"# Sheet: {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c) for c in row]
            if any(cells):
                out.append("\t".join(cells))
    return "\n".join(out)


def _zip_to_text(path: str) -> str:
    """List a .zip's members and inline the text-like ones (xml/csv/json/txt/...)."""
    import zipfile
    text_ext = {".xml", ".txt", ".csv", ".tsv", ".json", ".jsonld", ".md",
                ".html", ".htm", ".rels", ".log"}
    out = []
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        out.append("ZIP contents: " + ", ".join(names))
        for n in names:
            if os.path.splitext(n)[1].lower() in text_ext:
                try:
                    out.append(f"\n# {n}\n" + z.read(n).decode("utf-8", "replace"))
                except Exception as e:
                    out.append(f"\n# {n} (unreadable: {e!r})")
    return "\n".join(out)


def read_file(path: str) -> str:
    """Read a local attached file and return its text content.

    Dispatches by extension: .xlsx (cells as tab-separated rows), .csv/.tsv, .zip
    (text members inlined), .pdf (extracted text), and .json/.jsonld/.xml/.txt/.md/
    .html (raw text). The task tells you the file's path; pass it here exactly.
    """
    if not path or not os.path.exists(path):
        return f"ERROR: file not found: {path!r}"
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xlsx":
            text = _xlsx_to_text(path)
        elif ext == ".zip":
            text = _zip_to_text(path)
        elif ext == ".pdf":
            with open(path, "rb") as f:
                text = _pdf_to_text(f.read())
        else:                                        # csv/tsv/json/jsonld/xml/txt/... -> raw text
            with open(path, encoding="utf-8", errors="replace") as f:
                text = f.read()
    except Exception as e:
        return f"ERROR: read_file failed for {path!r}: {e!r}"
    if len(text) > READ_FILE_MAX_CHARS:
        dropped = len(text) - READ_FILE_MAX_CHARS
        text = (text[:READ_FILE_MAX_CHARS] +
                f"\n\n[...{dropped} chars NOT shown — this is a PARTIAL view of the file. "
                f"To process the COMPLETE file (all rows/records), do NOT rely on the text "
                f"above; instead use run_python to open this path directly and compute over "
                f"all of it: {path}]")
    return text or f"(no readable text in {path})"


# ---- registry: name -> (callable, single string arg, OpenAI schema) ----------
_SPEC = {
    "web_search": (web_search, "query", "Search the web; returns title/URL/snippet results."),
    "fetch_url":  (fetch_url,  "url",   "Fetch a web page and return its readable text."),
    "run_python": (run_python, "code",  "Execute a self-contained Python 3 snippet; returns stdout."),
    "read_file":  (read_file,  "path",  "Read a local attached file (xlsx/csv/zip/json/xml/pdf/text); returns its text."),
}

TOOL_FUNCS = {name: fn for name, (fn, _arg, _desc) in _SPEC.items()}

# Per-benchmark tool profiles. Add a benchmark -> pick a subset; the pipeline code
# never assumes any particular profile.
TOOL_PROFILES = {
    "none":        [],
    "math":        ["run_python"],
    "web":         ["web_search", "fetch_url"],
    "web_compute": ["web_search", "fetch_url", "run_python", "read_file"],
}


def tool_specs(names):
    """OpenAI `tools=` schemas for the given tool names ([] -> None)."""
    specs = []
    for n in names:
        fn, arg, desc = _SPEC[n]
        specs.append(dict(type="function", function=dict(
            name=n, description=(fn.__doc__ or desc).strip().split("\n")[0],
            parameters=dict(type="object",
                            properties={arg: dict(type="string", description=desc)},
                            required=[arg]))))
    return specs or None


def run_tool(name: str, arguments) -> str:
    """Dispatch one structured tool call. `arguments` is the OpenAI JSON-string
    (or already a dict). Pulls the single declared arg and calls the function."""
    if name not in _SPEC:
        return f"ERROR: unknown tool {name!r}."
    fn, arg, _desc = _SPEC[name]
    try:
        a = arguments if isinstance(arguments, dict) else json.loads(arguments or "{}")
    except Exception:
        a = {}
    val = a.get(arg)
    if val is None and len(a) == 1:          # tolerate a differently-named single key
        val = next(iter(a.values()))
    if not isinstance(val, str):
        val = "" if val is None else str(val)
    try:
        return fn(val)
    except Exception as e:
        return f"ERROR: tool {name} raised {e!r}"


if __name__ == "__main__":
    print(run_python("print(18/100*2350 + 17)"))
    print("---", tool_specs(["run_python", "web_search"]))
